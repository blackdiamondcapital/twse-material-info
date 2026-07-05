from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.config import MARKET_LABELS

COLUMNS = [
    ("market", "市場"),
    ("stock_code", "公司代號"),
    ("company_name", "公司名稱"),
    ("report_date", "出表日期"),
    ("announce_date", "發言日期"),
    ("announce_time", "發言時間"),
    ("subject", "主旨"),
    ("clause", "符合條款"),
    ("event_date", "事實發生日"),
    ("description", "說明"),
]


def build_excel(records: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "重大訊息"

    header_font = Font(bold=True)
    for col_idx, (_, label) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font

    for row_idx, rec in enumerate(records, start=2):
        for col_idx, (key, _) in enumerate(COLUMNS, start=1):
            value = rec.get(key, "")
            if key == "market":
                value = MARKET_LABELS.get(value, value)
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_idx, (key, _) in enumerate(COLUMNS, start=1):
        max_len = len(COLUMNS[col_idx - 1][1])
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            if row[0].value:
                max_len = max(max_len, min(len(str(row[0].value)), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
